from __future__ import annotations

import csv
import re
from collections.abc import Iterator
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from conferencia.domain.box_codes import content_hash_caixa_estoque, normalize_caixa_estoque


class ExcelReadError(Exception):
    """Validation error returned to the upload endpoint."""


class PalletFileImporter:
    """The sole CSV/XLSX interpretation path for pallet uploads."""

    ACCEPTED_COLUMNS = frozenset({
        "CAIXA_ESTOQUE", "CAIXA_ESTOQUE VARCHAR2", "CAIXA ESTOQUE",
    })
    CLASS_COLUMNS = frozenset({"DS_CLASSE", "DS CLASSE", "DS_CLASSE VARCHAR2"})

    def read_carton_codes(self, file_path: Path, extension: str) -> list[str]:
        return [item["caixa_estoque"] for item in self.read_expected_items(file_path, extension)]

    def read_expected_items(self, file_path: Path, extension: str) -> list[dict[str, str]]:
        rows = self._rows(file_path, extension)
        try:
            try:
                header = next(rows)
            except StopIteration as error:
                raise ExcelReadError("O arquivo nao possui cabecalho.") from error
            index = self._required_column_index(header)
            class_index = self._optional_column_index(header, self.CLASS_COLUMNS)
            items: list[dict[str, str]] = []
            for line_number, row in enumerate(rows, start=2):
                if index >= len(row):
                    continue
                code = self._cell_value(row[index], line_number)
                if code:
                    classe = self._class_value(row[class_index]) if class_index is not None and class_index < len(row) else ""
                    items.append({"caixa_estoque": code, "ds_classe": classe})
            if not items:
                raise ExcelReadError("Nao ha caixas validas na coluna CAIXA_ESTOQUE.")
            return items
        finally:
            close = getattr(rows, "close", None)
            if close is not None:
                close()

    @staticmethod
    def fingerprint(codes: list[object]) -> str:
        return content_hash_caixa_estoque(codes)

    def _rows(self, path: Path, extension: str) -> Iterator[list[object]]:
        if extension.casefold() == ".csv":
            return self._csv_rows(path)
        if extension.casefold() == ".xlsx":
            return self._xlsx_rows(path)
        raise ExcelReadError("Formato invalido. Envie um arquivo .csv ou .xlsx.")

    def _xlsx_rows(self, path: Path) -> Iterator[list[object]]:
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except (InvalidFileException, OSError, ValueError) as error:
            raise ExcelReadError("Nao foi possivel ler o arquivo .xlsx.") from error
        try:
            for row in workbook.active.iter_rows():
                yield list(row)
        finally:
            workbook.close()

    def _csv_rows(self, path: Path) -> Iterator[list[str]]:
        try:
            raw = path.read_text(encoding="utf-8-sig")
        except (OSError, UnicodeError) as error:
            raise ExcelReadError("O CSV deve estar codificado em UTF-8.") from error
        try:
            dialect = csv.Sniffer().sniff(raw[:8192], delimiters=";,")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";" if raw.partition("\n")[0].count(";") else ","
        try:
            yield from csv.reader(raw.splitlines(), dialect)
        except csv.Error as error:
            raise ExcelReadError("Nao foi possivel ler o arquivo .csv.") from error

    def _required_column_index(self, header: list[object]) -> int:
        for index, cell in enumerate(header):
            if self._normalize_header(self._raw_value(cell)) in self.ACCEPTED_COLUMNS:
                return index
        received = [self._raw_value(cell) for cell in header]
        raise ExcelReadError(
            "A coluna obrigat\u00f3ria CAIXA_ESTOQUE n\u00e3o foi encontrada. "
            f"Cabe\u00e7alhos recebidos: {received!r}"
        )

    def _optional_column_index(self, header: list[object], accepted: frozenset[str]) -> int | None:
        return next((index for index, cell in enumerate(header) if self._normalize_header(self._raw_value(cell)) in accepted), None)

    @staticmethod
    def _class_value(cell: object) -> str:
        return normalize_caixa_estoque(cell.value if hasattr(cell, "value") else cell).upper()

    @staticmethod
    def _normalize_header(value: str) -> str:
        return re.sub(r"\s+", " ", value.strip().upper())

    @staticmethod
    def _raw_value(cell: object) -> str:
        return normalize_caixa_estoque(cell.value if hasattr(cell, "value") else cell)

    @staticmethod
    def _cell_value(cell: object, line_number: int) -> str:
        value = cell.value if hasattr(cell, "value") else cell
        if value is None:
            return ""
        if isinstance(value, str):
            return normalize_caixa_estoque(value)
        if isinstance(value, (int, float)):
            if len(re.sub(r"\D", "", str(value))) >= 15:
                raise ExcelReadError(
                    "O arquivo possui codigos longos armazenados como numero no Excel. "
                    "Alguns digitos podem ter sido alterados. Exporte novamente o relatorio "
                    "mantendo CAIXA_ESTOQUE como texto."
                )
            raise ExcelReadError(
                f"Linha {line_number}: CAIXA_ESTOQUE esta armazenada como numero. "
                "Exporte CAIXA_ESTOQUE como texto para preservar o codigo."
            )
        return normalize_caixa_estoque(value)
